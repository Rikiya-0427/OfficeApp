from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active
ws2 = wb.create_sheet()

# Data can be assigned directly to cells
ws['A1'] = 42

# 数字を1から100まで入力する
for i in range(1, 101):
    row = ((i - 1) // 10) + 1  # 行の計算
    column = ((i - 1) % 10) + 1  # 列の計算
    ws2.cell(row=row, column=column).value = i

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")